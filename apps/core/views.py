from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import timedelta, datetime
import json

# Import models with proper app references
try:
    from orders.models import Order, OrderItem
    from products.models import Product, Category
    from users.models import CustomUser
    from payment.models import PaymentTransaction
    from core.models import ContactMessage
    from core.forms import ContactForm
except ImportError:
    # Fallback imports for when running from different contexts
    from apps.orders.models import Order, OrderItem
    from apps.products.models import Product, Category
    from apps.users.models import CustomUser
    from apps.payment.models import PaymentTransaction
    from apps.core.models import ContactMessage
    from apps.core.forms import ContactForm

def home(request):
    """Home page view"""
    # Fetch featured products to display on home page
    featured_products = Product.objects.filter(
        is_featured=True, 
        available=True
    ).select_related('category').order_by('-created_at')[:8]  # Limit to 8 featured products
    
    context = {
        'featured_products': featured_products,
    }
    return render(request, 'core/home.html', context)

def about(request):
    """About page view"""
    return render(request, 'core/about.html')

def contact(request):
    """Contact page view"""
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            try:
                # Get client IP and user agent
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    ip = x_forwarded_for.split(',')[0]
                else:
                    ip = request.META.get('REMOTE_ADDR')
                
                contact_message = form.save(commit=False)
                contact_message.ip_address = ip
                contact_message.user_agent = request.META.get('HTTP_USER_AGENT', '')
                contact_message.save()
                
                # Send automated confirmation email
                from core.email_utils import send_contact_confirmation_email
                try:
                    send_contact_confirmation_email(contact_message)
                except Exception as email_error:
                    # Log email error but don't fail the contact form submission
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Failed to send confirmation email: {email_error}")
                
                # Return success response
                return render(request, 'core/contact.html', {
                    'form': ContactForm(),  # Fresh form
                    'success': True,
                    'message': 'Thank you for your message! We will get back to you soon.'
                })
            except Exception as e:
                # Log the error and show user-friendly message
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error saving contact message: {e}")
                
                return render(request, 'core/contact.html', {
                    'form': form,
                    'error': True,
                    'message': 'Sorry, there was an error sending your message. Please try again.'
                })
        else:
            # Form has validation errors
            return render(request, 'core/contact.html', {
                'form': form,
                'error': True,
                'message': 'Please correct the errors below and try again.'
            })
    else:
        form = ContactForm()
    
    return render(request, 'core/contact.html', {'form': form})

def clear_notification(request):
    """Clear notification view - clears session notifications and marks contact messages as read"""
    if request.method == 'POST':
        try:
            # Clear session notification if it exists
            notification_cleared = False
            if 'notification' in request.session:
                del request.session['notification']
                notification_cleared = True
            
            # Also mark any unread contact messages as read (admin functionality)
            if request.user.is_authenticated and request.user.is_staff:
                updated_count = ContactMessage.objects.filter(status='new').update(status='read')
                if updated_count > 0:
                    return JsonResponse({
                        'success': True, 
                        'message': f'Cleared session notification and marked {updated_count} contact messages as read',
                        'notification_cleared': notification_cleared,
                        'contact_messages_updated': updated_count
                    })
            
            if notification_cleared:
                return JsonResponse({
                    'success': True, 
                    'message': 'Session notification cleared',
                    'notification_cleared': True,
                    'contact_messages_updated': 0
                })
            else:
                return JsonResponse({
                    'success': True, 
                    'message': 'No notification to clear',
                    'notification_cleared': False,
                    'contact_messages_updated': 0
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Error clearing notifications: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

def dashboard_callback(request, context):
    """Dashboard callback for admin interface - returns context data for admin dashboard"""
    # This function should return a dictionary of context data for the admin dashboard
    # It's called by Django Unfold to provide additional context
    try:
        # Get basic statistics for the dashboard
        total_orders = Order.objects.count()
        pending_orders = Order.objects.filter(status='pending').count()
        total_customers = CustomUser.objects.filter(user_type='customer').count()
        
        # Get contact message statistics
        total_contact_messages = ContactMessage.objects.count()
        new_contact_messages = ContactMessage.objects.filter(status='new').count()
        unread_contact_messages = ContactMessage.objects.filter(status__in=['new', 'read']).count()
        
        # Update the existing context with our custom data
        context.update({
            'custom_dashboard': True,
            'bakery_name': 'Live Bakery',
            'location': 'Kamalbinayak, Bhaktapur',
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'total_customers': total_customers,
            'total_contact_messages': total_contact_messages,
            'new_contact_messages': new_contact_messages,
            'unread_contact_messages': unread_contact_messages,
        })
        
        return context
        
    except Exception as e:
        # Return minimal context if there's an error
        context.update({
            'custom_dashboard': True,
            'bakery_name': 'Live Bakery',
            'location': 'Kamalbinayak, Bhaktapur',
            'error': str(e)
        })
        return context

@staff_member_required
def dashboard_api(request):
    """API endpoint for dashboard data"""
    try:
        # Get time period filter (default to current month)
        time_period = request.GET.get('period', 'current_month')
        selected_month = request.GET.get('month', None)  # Format: YYYY-MM
        selected_year = request.GET.get('year', None)    # Format: YYYY
        
        # Calculate date range based on period
        end_date = timezone.now()
        
        if time_period == 'yearly' and selected_year:
            # Specific year data
            start_date = timezone.datetime(int(selected_year), 1, 1, tzinfo=timezone.get_current_timezone())
            end_date = timezone.datetime(int(selected_year), 12, 31, 23, 59, 59, tzinfo=timezone.get_current_timezone())
        elif time_period == 'monthly' and selected_month:
            # Specific month data (format: YYYY-MM)
            year, month = map(int, selected_month.split('-'))
            start_date = timezone.datetime(year, month, 1, tzinfo=timezone.get_current_timezone())
            # Get last day of the month
            if month == 12:
                end_date = timezone.datetime(year, 12, 31, 23, 59, 59, tzinfo=timezone.get_current_timezone())
            else:
                end_date = timezone.datetime(year, month + 1, 1, tzinfo=timezone.get_current_timezone()) - timedelta(seconds=1)
        else:
            # Default: Current month
            start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            # End date is now
            end_date = timezone.now()
        
        # Basic statistics (filtered by time period)
        total_orders = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        total_revenue = Order.objects.filter(
            payment_status='paid',
            created_at__gte=start_date,
            created_at__lte=end_date
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        pending_orders = Order.objects.filter(status='pending').count()
        total_products = Product.objects.filter(available=True).count()
        
        # Total customers - cumulative count up to the end of selected period
        # This includes all customers registered from the beginning up to the selected month/year
        total_customers = CustomUser.objects.filter(
            user_type='customer',
            date_joined__lte=end_date
        ).count()
        
        # Contact message statistics (filtered by time period)
        total_contact_messages = ContactMessage.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        new_contact_messages = ContactMessage.objects.filter(
            status='new',
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        unread_contact_messages = ContactMessage.objects.filter(
            status__in=['new', 'read'],
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        replied_contact_messages = ContactMessage.objects.filter(
            status='replied',
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        # Cake orders (orders with cake items) - filtered by time period
        cake_orders = OrderItem.objects.filter(
            product__is_cake=True,
            order__created_at__gte=start_date,
            order__created_at__lte=end_date
        ).values('order').distinct().count()
        
        # Revenue data based on period
        daily_revenue = {}
        if time_period == 'yearly':
            # Group by month for yearly view
            for month_num in range(1, 13):
                month_start = timezone.datetime(int(selected_year or end_date.year), month_num, 1, tzinfo=timezone.get_current_timezone())
                if month_num == 12:
                    month_end = timezone.datetime(int(selected_year or end_date.year), 12, 31, 23, 59, 59, tzinfo=timezone.get_current_timezone())
                else:
                    month_end = timezone.datetime(int(selected_year or end_date.year), month_num + 1, 1, tzinfo=timezone.get_current_timezone()) - timedelta(seconds=1)
                
                month_str = month_start.strftime('%Y-%m')
                
                revenue = Order.objects.filter(
                    payment_status='paid',
                    created_at__gte=month_start,
                    created_at__lte=month_end
                ).aggregate(total=Sum('total_amount'))['total'] or 0
                daily_revenue[month_str] = float(revenue)
        else:
            # Daily revenue for monthly view
            current_date = start_date.date()
            end_date_only = end_date.date()
            
            while current_date <= end_date_only:
                date_str = current_date.strftime('%Y-%m-%d')
                revenue = Order.objects.filter(
                    payment_status='paid',
                    created_at__date=current_date
                ).aggregate(total=Sum('total_amount'))['total'] or 0
                daily_revenue[date_str] = float(revenue)
                current_date += timedelta(days=1)
        
        # Order status distribution (filtered by time period)
        order_status = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).values('status').annotate(count=Count('id'))
        order_status_dict = {item['status']: item['count'] for item in order_status}
        
        # Payment methods distribution (filtered by time period)
        payment_methods = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).values('payment_method').annotate(count=Count('id'))
        payment_methods_dict = {item['payment_method']: item['count'] for item in payment_methods}
        
        # Delivery types distribution (filtered by time period)
        delivery_types = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).values('delivery_type').annotate(count=Count('id'))
        delivery_types_dict = {item['delivery_type']: item['count'] for item in delivery_types}
        
        # Top selling products (filtered by time period)
        top_products = OrderItem.objects.filter(
            order__payment_status='paid',
            order__created_at__gte=start_date,
            order__created_at__lte=end_date
        ).values('product__name').annotate(
            quantity_sold=Sum('quantity'),
            revenue=Sum(F('quantity') * F('price'))
        ).order_by('-quantity_sold')[:10]
        
        top_products_list = [
            {
                'product_name': item['product__name'],
                'quantity_sold': item['quantity_sold'],
                'revenue': float(item['revenue'])
            }
            for item in top_products
        ]
        
        # Category performance (filtered by time period)
        category_performance = OrderItem.objects.filter(
            order__payment_status='paid',
            order__created_at__gte=start_date,
            order__created_at__lte=end_date
        ).values('product__category__name').annotate(
            quantity_sold=Sum('quantity'),
            revenue=Sum(F('quantity') * F('price'))
        ).order_by('-revenue')[:10]
        
        category_performance_list = [
            {
                'category_name': item['product__category__name'],
                'quantity_sold': item['quantity_sold'],
                'revenue': float(item['revenue'])
            }
            for item in category_performance
        ]
        
        # Recent orders
        recent_orders = Order.objects.select_related('user').prefetch_related(
            'items__product'
        ).order_by('-created_at')[:10]
        
        recent_orders_list = []
        for order in recent_orders:
            user_initials = ''
            if order.user.first_name and order.user.last_name:
                user_initials = f"{order.user.first_name[0]}{order.user.last_name[0]}"
            elif order.user.first_name:
                user_initials = order.user.first_name[0]
            elif order.user.username:
                user_initials = order.user.username[0].upper()
            
            recent_orders_list.append({
                'id': order.id,
                'order_number': order.order_number,
                'user_name': f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username,
                'user_initials': user_initials,
                'total_amount': float(order.total_amount),
                'status': order.status,
                'status_display': order.get_status_display(),
                'payment_status': order.payment_status,
                'payment_status_display': order.get_payment_status_display(),
                'delivery_type': order.delivery_type,
                'created_at': order.created_at.isoformat(),
                'item_count': order.items.count(),
            })
        
        # Recent contact messages
        recent_contact_messages = ContactMessage.objects.order_by('-created_at')[:10]
        
        recent_contact_list = []
        for contact in recent_contact_messages:
            recent_contact_list.append({
                'id': contact.id,
                'full_name': contact.full_name,
                'email': contact.email,
                'subject': contact.subject,
                'subject_display': contact.get_subject_display(),
                'status': contact.status,
                'status_display': contact.get_status_display(),
                'message': contact.message[:100] + '...' if len(contact.message) > 100 else contact.message,
                'created_at': contact.created_at.isoformat(),
                'days_since_creation': contact.get_days_since_creation(),
            })
        
        # Compile response data
        data = {
            'time_period': time_period,
            'selected_month': selected_month,
            'selected_year': selected_year,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'total_revenue': float(total_revenue),
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'total_products': total_products,
            'total_customers': total_customers,
            'cake_orders': cake_orders,
            'total_contact_messages': total_contact_messages,
            'new_contact_messages': new_contact_messages,
            'unread_contact_messages': unread_contact_messages,
            'replied_contact_messages': replied_contact_messages,
            'daily_revenue': daily_revenue,
            'order_status': order_status_dict,
            'payment_methods': payment_methods_dict,
            'delivery_types': delivery_types_dict,
            'top_products': top_products_list,
            'category_performance': category_performance_list,
            'recent_orders': recent_orders_list,
            'recent_contact_messages': recent_contact_list,
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'message': 'Failed to fetch dashboard data'
        }, status=500)


@staff_member_required
def dashboard_export_excel(request):
    """Export dashboard data to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get time period filter
        time_period = request.GET.get('period', 'monthly')
        selected_month = request.GET.get('month', None)
        selected_year = request.GET.get('year', None)
        
        # Calculate date range
        end_date = timezone.now()
        
        if time_period == 'yearly' and selected_year:
            start_date = timezone.datetime(int(selected_year), 1, 1, tzinfo=timezone.get_current_timezone())
            end_date = timezone.datetime(int(selected_year), 12, 31, 23, 59, 59, tzinfo=timezone.get_current_timezone())
            period_label = f"Year {selected_year}"
        elif time_period == 'monthly' and selected_month:
            year, month = map(int, selected_month.split('-'))
            start_date = timezone.datetime(year, month, 1, tzinfo=timezone.get_current_timezone())
            if month == 12:
                end_date = timezone.datetime(year, 12, 31, 23, 59, 59, tzinfo=timezone.get_current_timezone())
            else:
                end_date = timezone.datetime(year, month + 1, 1, tzinfo=timezone.get_current_timezone()) - timedelta(seconds=1)
            month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                          'July', 'August', 'September', 'October', 'November', 'December']
            period_label = f"{month_names[month - 1]} {year}"
        else:
            start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            period_label = "Current Month"
        
        # Fetch data (reuse logic from dashboard_api)
        total_orders = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        total_revenue = Order.objects.filter(
            payment_status='paid',
            created_at__gte=start_date,
            created_at__lte=end_date
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        pending_orders = Order.objects.filter(status='pending').count()
        total_products = Product.objects.filter(available=True).count()
        total_customers = CustomUser.objects.filter(
            user_type='customer',
            date_joined__lte=end_date
        ).count()
        
        cake_orders = OrderItem.objects.filter(
            product__is_cake=True,
            order__created_at__gte=start_date,
            order__created_at__lte=end_date
        ).values('order').distinct().count()
        
        total_contact_messages = ContactMessage.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        new_contact_messages = ContactMessage.objects.filter(
            status='new',
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "Live Bakery - Dashboard Report"
        ws['A1'].font = Font(bold=True, size=16, color="D97706")
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Period: {period_label}"
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:D2')
        
        ws['A3'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(size=10, italic=True)
        ws.merge_cells('A3:D3')
        
        # Summary Statistics
        row = 5
        ws[f'A{row}'] = "Summary Statistics"
        ws[f'A{row}'].font = title_font
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        
        stats = [
            ("Total Revenue", f"Rs. {total_revenue:,.2f}"),
            ("Total Orders", total_orders),
            ("Pending Orders", pending_orders),
            ("Active Products", total_products),
            ("Total Customers", total_customers),
            ("Cake Orders", cake_orders),
            ("Contact Messages", total_contact_messages),
            ("New Messages", new_contact_messages),
        ]
        
        for metric, value in stats:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
        
        # Order Status Distribution
        row += 3
        ws[f'A{row}'] = "Order Status Distribution"
        ws[f'A{row}'].font = title_font
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Status"
        ws[f'B{row}'] = "Count"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        
        order_status = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).values('status').annotate(count=Count('id'))
        
        for item in order_status:
            row += 1
            ws[f'A{row}'] = item['status'].capitalize()
            ws[f'B{row}'] = item['count']
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
        
        # Top Products
        row += 3
        ws[f'A{row}'] = "Top Selling Products"
        ws[f'A{row}'].font = title_font
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "Product"
        ws[f'B{row}'] = "Quantity Sold"
        ws[f'C{row}'] = "Revenue"
        ws[f'D{row}'] = "Orders"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
        
        top_products = OrderItem.objects.filter(
            order__payment_status='paid',
            order__created_at__gte=start_date,
            order__created_at__lte=end_date
        ).values('product__name').annotate(
            quantity_sold=Sum('quantity'),
            revenue=Sum(F('quantity') * F('price')),
            orders_count=Count('order', distinct=True)
        ).order_by('-quantity_sold')[:10]
        
        for item in top_products:
            row += 1
            ws[f'A{row}'] = item['product__name']
            ws[f'B{row}'] = item['quantity_sold']
            ws[f'C{row}'] = f"Rs. {item['revenue']:,.2f}"
            ws[f'D{row}'] = item['orders_count']
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Live_Bakery_Dashboard_{period_label.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting dashboard to Excel: {e}")
        return HttpResponse(f'Error generating Excel file: {str(e)}', status=500)